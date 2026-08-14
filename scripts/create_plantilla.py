#!/usr/bin/env python3
"""Genera plantilla Excel para carga de propiedades en Valu."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Columns in order matching valu_forms.py
COLUMNS = [
    ("nombre", "text", None),
    ("tipo_inmueble", "text", "departamento,casa,ph,local,oficina,terreno"),
    ("zona", "text", "Centro,Centro Premium,Norte,Oeste,Sur,Fisherton,Puerto Norte,Macrocentro,Resto"),
    ("direccion", "text", None),
    ("lat", "number", None),
    ("lon", "number", None),
    ("ubicacion_tipo", "text", "calle,pasaje,avenida"),
    ("m2_cubiertos", "number", None),
    ("m2_semicubiertos", "number", None),
    ("m2_descubiertos_propios", "number", None),
    ("m2_descubiertos_comun_exclusivo", "number", None),
    ("dormitorios", "number", None),
    ("ambientes", "number", None),
    ("baños", "number", None),
    ("toilet", "text", "TRUE,FALSE"),
    ("baño_servicio", "text", "TRUE,FALSE"),
    ("anio_construccion", "number", None),
    ("constructora", "text", None),
    ("piso", "number", None),
    ("total_pisos", "number", None),
    ("vista", "text", "frente,contrafrente,lateral,libre"),
    ("disposicion", "text", "pasante,contrafrente,interna,frente"),
    ("gas_ok", "text", "si,no,en_proceso"),
    ("tipo_balcon", "text", "ninguno,corrido,L,frances,terraza,terraza_servicio"),
    ("orientacion", "text", "norte,sur,este,oeste,noreste,noroeste,sureste,suroeste"),
    ("ventilacion", "text", "cruzada,simple"),
    ("estado_detalle", "text", "a_estrenar,bueno,regular,a_reciclar"),
    ("calidad_edificio", "text", "premium,media,economica"),
    ("seguridad", "text", "ninguna,tag,camaras,24hs,totem"),
    ("terminaciones_suelo", "text", "madera,porcelanato,ceramico,vinilico,estandar (separar con coma si varios)"),
    ("carpinteria", "text", "piso_techo,dvh,estandar"),
    ("terminaciones_cocina", "text", "silestone,granito,estandar"),
    ("preinstalacion_aa", "text", "TRUE,FALSE"),
    ("cocheras_cantidad", "number", None),
    ("cocheras_tipo", "text", "cubierta,descubierta,semicubierta"),
    ("valor_cochera_base", "number", None),
    ("valor_baulera", "number", None),
    ("doble_ingreso", "text", "TRUE,FALSE"),
    ("lavadero_independiente", "text", "TRUE,FALSE"),
    ("reciclado", "text", "TRUE,FALSE"),
    ("reciclado_tipo", "text", "ninguno,parcial,total"),
    ("anio_reciclado", "number", None),
    ("ventilacion_bano", "text", "natural,forzada,sin_ventana"),
    ("layout_flexible", "text", "TRUE,FALSE"),
    ("placares_completos", "text", "TRUE,FALSE"),
    ("despensa", "text", "TRUE,FALSE"),
    ("ascensores_edificio", "number", None),
    ("detalles_categoria", "text", "caldera_central,radiadores,seguridad_24hs,seguridad_tag,seguridad_camaras,seguridad_totem,aberturas_premium,parrilla_propia,parrilla_compartida,terraza_compartida,pileta,sum,gym,quincho,marinas,co_working,esparcimiento (separar con coma si varios)"),
    ("descripcion_libre", "text", None),
    ("valor_compra_usd", "number", None),
    ("fecha_compra", "text", None),
    ("fecha_publicacion", "text", None),
    ("expensas_ars", "number", None),
    ("id", "text", None),
]

# Example data (Cochabamba 45)
EXAMPLE = {
    "nombre": "Cochabamba 45",
    "tipo_inmueble": "departamento",
    "zona": "Macrocentro",
    "direccion": "Cochabamba 45",
    "lat": -32.9524509,
    "lon": -60.6343631,
    "ubicacion_tipo": "calle",
    "m2_cubiertos": 98,
    "m2_semicubiertos": 0,
    "m2_descubiertos_propios": 0,
    "m2_descubiertos_comun_exclusivo": 0,
    "dormitorios": 4,
    "ambientes": 4,
    "baños": 2,
    "toilet": "FALSE",
    "baño_servicio": "FALSE",
    "anio_construccion": 1970,
    "constructora": "",
    "piso": 2,
    "total_pisos": 4,
    "vista": "frente",
    "disposicion": "pasante",
    "gas_ok": "si",
    "tipo_balcon": "ninguno",
    "orientacion": "este",
    "ventilacion": "cruzada",
    "estado_detalle": "regular",
    "calidad_edificio": "media",
    "seguridad": "ninguna",
    "terminaciones_suelo": "porcelanato",
    "carpinteria": "estandar",
    "terminaciones_cocina": "estandar",
    "preinstalacion_aa": "FALSE",
    "cocheras_cantidad": 0,
    "cocheras_tipo": "cubierta",
    "valor_cochera_base": 0,
    "valor_baulera": 0,
    "doble_ingreso": "FALSE",
    "lavadero_independiente": "FALSE",
    "reciclado": "FALSE",
    "reciclado_tipo": "ninguno",
    "anio_reciclado": "",
    "ventilacion_bano": "natural",
    "layout_flexible": "FALSE",
    "placares_completos": "FALSE",
    "despensa": "TRUE",
    "ascensores_edificio": 1,
    "detalles_categoria": "",
    "descripcion_libre": "Departamento pasante, piso 2, vista a Cochabamba. Constructura sólida, necesita reciclado parcial.",
    "valor_compra_usd": 0,
    "fecha_compra": "",
    "fecha_publicacion": "2026-07-31",
    "expensas_ars": 0,
    "id": "",
}

# Multi-select fields
MULTI_FIELDS = {"terminaciones_suelo", "detalles_categoria"}

# Fields with dropdowns (only the ones with short lists)
DROPDOWN_FIELDS = {
    "tipo_inmueble": "departamento,casa,ph,local,oficina,terreno",
    "zona": "Centro,Centro Premium,Norte,Oeste,Sur,Fisherton,Puerto Norte,Macrocentro,Resto",
    "ubicacion_tipo": "calle,pasaje,avenida",
    "vista": "frente,contrafrente,lateral,libre",
    "disposicion": "pasante,contrafrente,interna,frente",
    "gas_ok": "si,no,en_proceso",
    "tipo_balcon": "ninguno,corrido,L,frances,terraza,terraza_servicio",
    "orientacion": "norte,sur,este,oeste,noreste,noroeste,sureste,suroeste",
    "ventilacion": "cruzada,simple",
    "estado_detalle": "a_estrenar,bueno,regular,a_reciclar",
    "calidad_edificio": "premium,media,economica",
    "seguridad": "ninguna,tag,camaras,24hs,totem",
    "carpinteria": "piso_techo,dvh,estandar",
    "terminaciones_cocina": "silestone,granito,estandar",
    "cocheras_tipo": "cubierta,descubierta,semicubierta",
    "reciclado_tipo": "ninguno,parcial,total",
    "ventilacion_bano": "natural,forzada,sin_ventana",
    "toilet": "TRUE,FALSE",
    "baño_servicio": "TRUE,FALSE",
    "preinstalacion_aa": "TRUE,FALSE",
    "doble_ingreso": "TRUE,FALSE",
    "lavadero_independiente": "TRUE,FALSE",
    "reciclado": "TRUE,FALSE",
    "layout_flexible": "TRUE,FALSE",
    "placares_completos": "TRUE,FALSE",
    "despensa": "TRUE,FALSE",
}


def create_plantilla():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Propiedades ──
    ws = wb.active
    ws.title = "Propiedades"

    header_fill = PatternFill(start_color="006AFF", end_color="006AFF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    example_fill = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")

    # Write headers
    for col_idx, (name, _, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Write example row
    for col_idx, (name, _, _) in enumerate(COLUMNS, 1):
        val = EXAMPLE.get(name, "")
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.fill = example_fill

    # Add data validations
    for col_idx, (name, _, options) in enumerate(COLUMNS, 1):
        if name in DROPDOWN_FIELDS:
            formula = DROPDOWN_FIELDS[name]
            dv = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
            dv.error = f"Seleccione una opción válida: {formula}"
            dv.errorTitle = "Opción inválida"
            dv.prompt = f"Opciones: {formula}"
            dv.promptTitle = name
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            dv.add(f"{col_letter}2:{col_letter}1000")
            ws.add_data_validation(dv)

    # Column widths
    for col_idx, (name, _, _) in enumerate(COLUMNS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(name) + 2, 14)

    # Freeze header
    ws.freeze_panes = "A2"

    # ── Sheet 2: Instrucciones ──
    ws2 = wb.create_sheet("Instrucciones")
    instructions = [
        ("INSTRUCCIONES DE CARGA", "", True),
        ("", "", False),
        ("CAMPOS REQUERIDOS:", "", True),
        ("nombre", "Nombre único de la propiedad (ej: 'Cochabamba 45')", False),
        ("tipo_inmueble", "departamento | casa | ph | local | oficina | terreno", False),
        ("zona", "Centro | Centro Premium | Norte | Oeste | Sur | Fisherton | Puerto Norte | Macrocentro | Resto", False),
        ("direccion", "Dirección completa", False),
        ("lat", "Latitud GPS", False),
        ("lon", "Longitud GPS", False),
        ("m2_cubiertos", "Metros cuadrados cubiertos", False),
        ("dormitorios", "Cantidad de dormitorios", False),
        ("anio_construccion", "Año de construcción (ej: 1970)", False),
        ("", "", False),
        ("CAMPOS MULTI-SELECCIÓN (separar con coma):", "", True),
        ("terminaciones_suelo", "madera, porcelanato, ceramico, vinilico, estandar. Si tiene varios: 'madera,ceramico'", False),
        ("detalles_categoria", "caldera_central, radiadores, seguridad_24hs, seguridad_tag, seguridad_camaras, seguridad_totem, aberturas_premium, parrilla_propia, parrilla_compartida, terraza_compartida, pileta, sum, gym, quincho, marinas, co_working, esparcimiento", False),
        ("", "", False),
        ("CAMPOS TRUE/FALSE:", "", True),
        ("toilet, baño_servicio, preinstalacion_aa, doble_ingreso, lavadero_independiente, reciclado, layout_flexible, placares_completos, despensa", "Usar TRUE o FALSE", False),
        ("", "", False),
        ("CAMPOS FECHA:", "", True),
        ("fecha_compra, fecha_publicacion", "Formato: YYYY-MM-DD (ej: 2026-07-31)", False),
        ("", "", False),
        ("CAMPOS NUMÉRICOS OPCIONALES:", "", True),
        ("m2_semicubiertos, m2_descubiertos_propios, m2_descubiertos_comun_exclusivo", "Default: 0", False),
        ("ambientes, baños, piso, total_pisos, cocheras_cantidad, ascensores_edificio", "Números enteros", False),
        ("valor_cochera_base, valor_baulera, expensas_ars, valor_compra_usd", "Números decimales", False),
        ("", "", False),
        ("ID:", "", True),
        ("id", "Dejar vacío para auto-generar. Si se proporciona, debe ser único.", False),
    ]

    for row_idx, (field, desc, is_header) in enumerate(instructions, 1):
        c1 = ws2.cell(row=row_idx, column=1, value=field)
        c2 = ws2.cell(row=row_idx, column=2, value=desc)
        if is_header:
            c1.font = Font(bold=True, size=12)
    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 80

    # Save
    output_path = os.path.join(BASE_DIR, "plantilla_propiedades.xlsx")
    wb.save(output_path)
    print(f"Plantilla generada: {output_path}")
    return output_path


if __name__ == "__main__":
    create_plantilla()
