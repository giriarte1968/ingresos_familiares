#!/usr/bin/env python3
"""Agrega columna zona + fila Entre Rios a la plantilla."""
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PLANTILLA = os.path.join(BASE_DIR, "plantilla_propiedades.xlsx")

wb = openpyxl.load_workbook(PLANTILLA)
ws = wb["Propiedades"]

# ── 1. Insertar columna 'zona' en posición 3 (antes de 'direccion') ──
ws.insert_cols(3)
zona_cell = ws.cell(row=1, column=3, value="zona")
zona_cell.fill = PatternFill(start_color="006AFF", end_color="006AFF", fill_type="solid")
zona_cell.font = Font(bold=True, color="FFFFFF", size=11)
ws.column_dimensions["C"].width = 18

# Mover ejemplo de Cochabamba 45: columna D (era C) ahora queda en E
# shift fue automático con insert_cols

# Agregar zona al row 2 (ejemplo)
ws.cell(row=2, column=3, value="Macrocentro")

# Agregar validación dropdown para zona
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type="list", formula1='"Centro,Centro Premium,Norte,Oeste,Sur,Fisherton,Puerto Norte,Macrocentro,Resto"', allow_blank=True)
dv.error = "Seleccione una zona válida"
dv.errorTitle = "Zona inválida"
dv.add("C2:C1000")
ws.add_data_validation(dv)

# ── 2. Agregar fila Entre Rios 1372 (row 3) ──
# Mapeo del Excel fuente -> columnas plantilla
entre_rios = {
    1: "Entre Rios 1372",        # nombre
    2: "departamento",            # tipo_inmueble
    3: "Centro",                  # zona
    4: "Entre Rios 1372",        # direccion
    5: "calle",                   # ubicacion_tipo
    6: 73.24,                     # m2_cubiertos
    7: 8.88,                      # m2_semicubiertos
    8: 21.47,                     # m2_descubiertos_propios
    9: 0,                         # m2_descubiertos_comun_exclusivo
    10: 2,                        # dormitorios
    11: 2,                        # ambientes
    12: 1,                        # baños
    13: "TRUE",                   # toilet (tiene toilette)
    14: "FALSE",                  # baño_servicio
    15: 1965,                     # anio_construccion
    16: "",                       # constructora
    17: 2,                        # piso (2do piso)
    18: 0,                        # total_pisos (no especificado)
    19: "frente",                 # vista (a la calle)
    20: "frente",                 # disposicion
    21: "si",                     # gas_ok
    22: "terraza",                # tipo_balcon (Terraza: si, tender)
    23: "este",                   # orientacion
    24: "cruzada",                # ventilacion
    25: "regular",                # estado_detalle
    26: "media",                  # calidad_edificio (Buena -> media)
    27: "ninguna",                # seguridad
    28: "ceramico",               # terminaciones_suelo (Parques y ceramicos)
    29: "estandar",               # carpinteria
    30: "estandar",               # terminaciones_cocina
    31: "FALSE",                  # preinstalacion_aa
    32: 0,                        # cocheras_cantidad (no)
    33: "cubierta",               # cocheras_tipo
    34: 0,                        # valor_cochera_base
    35: 0,                        # valor_baulera
    36: "TRUE",                   # doble_ingreso (Ingresos: 2)
    37: "TRUE",                   # lavadero_independiente
    38: "FALSE",                  # reciclado (no)
    39: "ninguno",                # reciclado_tipo
    40: "",                       # anio_reciclado
    41: "natural",                # ventilacion_bano
    42: "FALSE",                  # layout_flexible
    43: "TRUE",                   # placares_completos
    44: "TRUE",                   # despensa
    45: 0,                        # ascensores_edificio (escalera)
    46: "",                       # detalles_categoria
    47: "Departamento tipo ph, 2do piso de escalera, construccion solida pero precisa reparaciones. Dormitorio secundario con ventilacion a aire luz, sector de guardado con baño de servicio, dos dormitorios comodos, uno a la calle living con salida al balcon y cocina con ventilacion al aire luz.",
    48: 0,                        # valor_compra_usd
    49: "",                       # fecha_compra
    50: "",                       # fecha_publicacion
    51: 0,                        # expensas_ars
}

for col, val in entre_rios.items():
    ws.cell(row=3, column=col, value=val)

# Guardar
wb.save(PLANTILLA)
print(f"Plantilla actualizada: {PLANTILLA}")
print(f"Fila 3 agregada: Entre Rios 1372")
